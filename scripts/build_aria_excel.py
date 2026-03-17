"""
Builds ARIA_Initiative_Framework.xlsx with 7 sheets.
Tools based on easiest-to-build agents from iMocha_Skills_Intelligence_Agents.xlsx
iMocha brand colours: #220133 (dark purple), #FD5A0F (orange)
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ─────────────────────────────────────────────────────────────
PURPLE      = "220133"
ORANGE      = "FD5A0F"
ORANGE_LITE = "FFF0EA"
ORANGE_MID  = "FDBB96"
PURPLE_LITE = "F4EFF6"
PURPLE_MID  = "EAE4EF"
WHITE       = "FFFFFF"
GREY        = "F9F7FB"
TEXT_DARK   = "220133"
TEXT_MID    = "553366"
TEXT_MUTED  = "9988AA"
GREEN       = "059669"
GREEN_LITE  = "ECFDF5"
AMBER       = "D97706"
AMBER_LITE  = "FEF3C7"
RED_LITE    = "FEE2E2"
RED         = "DC2626"

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=TEXT_DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(wrap=True, h="left", v="center"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def border_thin():
    s = Side(style="thin", color=PURPLE_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, cols, bg=PURPLE, fg=WHITE, size=10, height=22):
    ws.row_dimensions[row].height = height
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=fg, size=size)
        cell.alignment = align(wrap=True, h="center")
        cell.border = border_thin()

def write_section_title(ws, row, col, text, colspan=1, bg=ORANGE, fg=WHITE):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=11)
    cell.alignment = align(h="left")
    cell.border = border_thin()
    ws.row_dimensions[row].height = 20
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — BRIEF
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "① Brief"

ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "PROJECT ARIA  —  AI-Native Reach & Intelligence Activation"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=16)
c.alignment = align(h="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = "CEO Sign-Off Brief  ·  iMocha CEO's Office Initiative  ·  Owner: Growth PM  ·  March 2026"
c.fill = fill(ORANGE)
c.font = font(bold=False, color=WHITE, size=10, italic=True)
c.alignment = align(h="center")
ws1.row_dimensions[2].height = 18

ws1.append([])  # spacer row 3

write_section_title(ws1, 4, 1, "THE BIG BET", colspan=6)

ws1.merge_cells("A5:F5")
c = ws1["A5"]
c.value = (
    "Most B2B SaaS companies sell skills intelligence. We will demonstrate it — through a portfolio of free, "
    "AI-powered tools that let HR leaders experience iMocha's value before a single sales conversation happens. "
    "We chose the 7 easiest-to-build agents from iMocha's 23-agent architecture: all text-in / structured-output, "
    "no complex infra, shippable by 1 PM + 1 intern in 10-day sprints. Each tool creates a personalised 'aha moment' "
    "for a specific persona, captures intent, and feeds a qualified pipeline. The tools do the selling."
)
c.fill = fill(GREY)
c.font = font(color=TEXT_MID, size=10)
c.alignment = align(wrap=True, h="left")
c.border = border_thin()
ws1.row_dimensions[5].height = 56

ws1.merge_cells("A6:F6")
c = ws1["A6"]
c.value = "TARGET: 5,000 relevant personas experience an iMocha AI tool within 90 days."
c.fill = fill(ORANGE_LITE)
c.font = font(bold=True, color=ORANGE, size=11)
c.alignment = align(h="center")
c.border = border_thin()
ws1.row_dimensions[6].height = 22

ws1.append([])  # spacer row 7

write_section_title(ws1, 8, 1, "WHY THESE 7 AGENTS — BUILD SELECTION LOGIC", colspan=6)

ws1.merge_cells("A9:F9")
c = ws1["A9"]
c.value = (
    "From iMocha's 23-agent architecture, we selected the 7 where ALL of these are true:\n"
    "  (1) Input = text the user can paste in a browser  "
    "(2) Output = structured analysis an LLM can generate  "
    "(3) No real-time adaptive systems, webcam feeds, or complex ML infra needed  "
    "(4) Core value is deliverable with Claude + LangGraph + Next.js — the same stack the team already owns"
)
c.fill = fill(PURPLE_LITE)
c.font = font(color=TEXT_MID, size=10)
c.alignment = align(wrap=True, h="left")
c.border = border_thin()
ws1.row_dimensions[9].height = 52

# Agent selection table
tool_rows = [
    ("#16", "AI Automation Index",       "Skills Architecture", "Skill list + AI signals → automation risk score per skill",          "CHRO · Workforce Planning", "✅ LIVE"),
    ("#1",  "Resume Inference",          "Skills Inference",    "Resume text (paste/upload) → structured skill profile + proficiency", "TA Leader · Recruiter",    "Sprint 2"),
    ("#14", "Skill Normalization",       "Skills Architecture", "Raw skill strings from any JD or profile → canonical skill names",    "TA Leader · People Ops",   "Sprint 3"),
    ("#2",  "Certification Inference",   "Skills Inference",    "Certification names → implied skills with proficiency levels",        "L&D Head · Employee",      "Sprint 4"),
    ("#18", "Career Path",               "Skills Matching",     "Current skills + career goal → 2-3 AI-safe career paths",            "L&D Head · Employee",      "Sprint 5"),
    ("#19", "Learning Recommendation",   "Skills Matching",     "Skill gap → prioritised personalised learning path",                  "L&D Head",                 "Sprint 6"),
    ("#15", "Emerging Skills Detection", "Skills Architecture", "Role / domain → rising skills before any framework captures them",   "CHRO · Analyst",           "Sprint 7"),
]

ws1.row_dimensions[10].height = 18
for col, hdr in enumerate(["Agent #", "Tool Name", "Category", "What it does (input → output)", "Primary Persona", "Status"], 1):
    c = ws1.cell(row=10, column=col, value=hdr)
    c.fill = fill(PURPLE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

status_fills = {"✅ LIVE": GREEN_LITE, "Sprint 2": ORANGE_LITE, "Sprint 3": ORANGE_LITE,
                "Sprint 4": AMBER_LITE, "Sprint 5": AMBER_LITE, "Sprint 6": PURPLE_LITE, "Sprint 7": PURPLE_LITE}
status_colors = {"✅ LIVE": GREEN, "Sprint 2": ORANGE, "Sprint 3": ORANGE,
                 "Sprint 4": AMBER, "Sprint 5": AMBER, "Sprint 6": PURPLE, "Sprint 7": PURPLE}

for ri, (num, name, cat, desc, persona, status) in enumerate(tool_rows, 11):
    bg = WHITE if ri % 2 == 1 else GREY
    ws1.row_dimensions[ri].height = 30
    for col, val in enumerate([num, name, cat, desc, persona, status], 1):
        c = ws1.cell(row=ri, column=col, value=val)
        if col == 6:
            c.fill = fill(status_fills.get(val, WHITE))
            c.font = font(bold=True, color=status_colors.get(val, TEXT_DARK), size=10)
            c.alignment = align(h="center")
        elif col in (1, 2):
            c.fill = fill(PURPLE_LITE)
            c.font = font(bold=(col == 2), color=TEXT_DARK, size=10)
            c.alignment = align(h="center" if col == 1 else "left")
        else:
            c.fill = fill(bg)
            c.font = font(color=TEXT_DARK, size=10)
            c.alignment = align()
        c.border = border_thin()

ws1.append([])  # spacer row 18

write_section_title(ws1, 19, 1, "THE THREE LAYERS", colspan=6)

layers = [
    ("LAYER 1 — BUILD", "1 PM + 1 Technical Intern + AI",
     "Ship 1 new AI tool every 10 days — 7 tools in 60 days",
     "Claude · LangGraph · Next.js · Vercel", PURPLE_LITE, TEXT_DARK),
    ("LAYER 2 — REACH", "AI-powered distribution to 5,000 personas",
     "Personalised outreach at scale — one tool per message, always",
     "Clay · Apollo · LinkedIn AI · Claude for copy", ORANGE_LITE, ORANGE),
    ("LAYER 3 — CONVERT", "In-tool lead capture → Sales handoff pipeline",
     "Every tool trial becomes a qualified conversation",
     "HubSpot · Slack alerts · AI follow-up sequences", GREEN_LITE, GREEN),
]

ws1.row_dimensions[20].height = 18
for col, hdr in enumerate(["Layer", "Team", "Goal", "Stack"], 1):
    c = ws1.cell(row=20, column=col, value=hdr)
    c.fill = fill(PURPLE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

for i, (layer, team, goal, stack, bg, tc) in enumerate(layers, 21):
    ws1.row_dimensions[i].height = 30
    for col, val in enumerate([layer, team, goal, stack], 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=(col == 1), color=tc if col == 1 else TEXT_DARK, size=10)
        c.alignment = align()
        c.border = border_thin()

ws1.append([])  # spacer row 24

write_section_title(ws1, 25, 1, "THE ASK — Sign-Off on Three Things", colspan=6)

asks = [
    ("1  BUDGET",       "Approve $15,000 for the 90-day ARIA pilot"),
    ("2  ACCESS",       "Connect Growth PM to Sales & Marketing ops for HubSpot + Clay + Apollo access"),
    ("3  GTM ALIGNMENT", "30-minute session with Sales leadership to agree lead handoff SLA — AE contacts a 2-tool user within 24 hours"),
]
for i, (k, v) in enumerate(asks, 26):
    ws1.row_dimensions[i].height = 30
    kc = ws1.cell(row=i, column=1, value=k)
    kc.fill = fill(PURPLE_LITE)
    kc.font = font(bold=True, color=PURPLE, size=10)
    kc.alignment = align()
    kc.border = border_thin()
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    vc = ws1.cell(row=i, column=2, value=v)
    vc.fill = fill(WHITE)
    vc.font = font(color=TEXT_DARK, size=10)
    vc.alignment = align()
    vc.border = border_thin()

ws1.merge_cells("A29:F29")
c = ws1["A29"]
c.value = ("If this pilot hits 250 sales-accepted leads in 90 days, ARIA becomes a permanent motion — "
           "a compounding AI-native pipeline engine that scales with every new tool we ship.")
c.fill = fill(ORANGE)
c.font = font(bold=True, color=WHITE, size=10)
c.alignment = align(h="center")
c.border = border_thin()
ws1.row_dimensions[29].height = 28

ws1.merge_cells("A30:F30")
c = ws1["A30"]
c.value = "Project ARIA  ·  iMocha CEO's Office  ·  March 2026  ·  \"We build the tools. The tools build the pipeline.\""
c.fill = fill(GREY)
c.font = font(italic=True, color=TEXT_MUTED, size=9)
c.alignment = align(h="center")
ws1.row_dimensions[30].height = 16

for col, w in zip("ABCDEF", [10, 26, 18, 42, 22, 12]):
    ws1.column_dimensions[col].width = w
ws1.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RELEASE CALENDAR
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("② Release Calendar")

ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "PROJECT ARIA — Tool Release Calendar  ·  1 Tool Every 10 Days  ·  All 7 Tools Live by Day 60"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws2.row_dimensions[1].height = 30

headers = ["Day", "Launch Date", "Tool Name", "Agent from iMocha Architecture", "Primary Persona", "The 'Aha Moment'", "Status"]
ws2.row_dimensions[2].height = 22
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = fill(ORANGE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

tools = [
    (0,  "Mar 7",  "① AI Automation Index",
     "#16 AI Automation Index Agent  ·  Skills Architecture",
     "CHRO · HR Head · Workforce Planning",
     "Every task in this role has a live automation risk score — and a reskilling urgency signal",
     "✅ LIVE"),
    (10, "Mar 17", "② Resume Skill Extractor",
     "#1 Resume Inference Agent  ·  Skills Inference",
     "TA Leader · Recruiter",
     "Paste a resume and see every skill extracted, proficiency-rated, and confidence-scored in seconds — no manual review",
     "Sprint 2"),
    (20, "Mar 27", "③ Skill Normalizer",
     "#14 Skill Normalization Agent  ·  Skills Architecture",
     "TA Leader · People Ops · HR Analyst",
     "Paste 20 skill names from any JD or profile — get the canonical version of each, deduplicated and framework-mapped",
     "Sprint 3"),
    (30, "Apr 6",  "④ Certification Skill Mapper",
     "#2 Certification Inference Agent  ·  Skills Inference",
     "L&D Head · Employee · HR Generalist",
     "List your certifications — know exactly which skills you've proven, at what proficiency, and which ones are expiring",
     "Sprint 4"),
    (40, "Apr 16", "⑤ Career Path Intelligence",
     "#18 Career Path Agent  ·  Skills Matching",
     "L&D Head · Employee · Manager",
     "Paste your current skills and a target role — get 3 AI-safe career paths with skill milestones and automation risk per path",
     "Sprint 5"),
    (50, "Apr 26", "⑥ Learning Path Generator",
     "#19 Learning Recommendation Agent  ·  Skills Matching",
     "L&D Head · Employee",
     "Enter your skill gap and time budget — get a prioritised learning plan with content type, estimated skill gain, and ROI projection",
     "Sprint 6"),
    (60, "May 6",  "⑦ Emerging Skills Radar",
     "#15 Emerging Skills Detection Agent  ·  Skills Architecture",
     "CHRO · L&D Head · HR Tech Analyst",
     "Enter any role or domain — see which skills are rising in the wild before ESCO, O*NET, or any framework has caught them",
     "Sprint 7"),
]

status_styles = {
    "✅ LIVE":   (GREEN_LITE, GREEN),
    "Sprint 2":  (ORANGE_LITE, ORANGE),
    "Sprint 3":  (ORANGE_LITE, ORANGE),
    "Sprint 4":  (AMBER_LITE, AMBER),
    "Sprint 5":  (AMBER_LITE, AMBER),
    "Sprint 6":  (PURPLE_LITE, PURPLE),
    "Sprint 7":  (PURPLE_LITE, PURPLE),
}

for row_i, (day, date, name, agents, persona, aha, status) in enumerate(tools, 3):
    bg = WHITE if row_i % 2 == 1 else GREY
    ws2.row_dimensions[row_i].height = 52
    vals = [day, date, name, agents, persona, aha, status]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row_i, column=col, value=val)
        if col == 7:
            sbg, sfg = status_styles.get(val, (WHITE, TEXT_DARK))
            c.fill = fill(sbg)
            c.font = font(bold=True, color=sfg, size=10)
            c.alignment = align(h="center")
        elif col == 1:
            c.fill = fill(PURPLE_LITE)
            c.font = font(bold=True, color=PURPLE, size=14)
            c.alignment = align(h="center")
        elif col == 3:
            c.fill = fill(bg)
            c.font = font(bold=True, color=TEXT_DARK, size=10)
            c.alignment = align()
        else:
            c.fill = fill(bg)
            c.font = font(color=TEXT_DARK, size=10)
            c.alignment = align()
        c.border = border_thin()

for col, w in zip(range(1, 8), [6, 11, 24, 40, 22, 54, 12]):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PERSONA × TOOLS MATRIX
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("③ Persona × Tools")

ws3.merge_cells("A1:J1")
c = ws3["A1"]
c.value = "PROJECT ARIA — Persona × Tool Matrix  ·  Which Tools Each Persona Receives & In What Order"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws3.row_dimensions[1].height = 30

tool_short = [
    "①\nAI Auto\nIndex",
    "②\nResume\nExtractor",
    "③\nSkill\nNormalizer",
    "④\nCert\nMapper",
    "⑤\nCareer\nPath",
    "⑥\nLearning\nPath",
    "⑦\nEmerging\nSkills",
]

ws3.row_dimensions[2].height = 46
for col, val in enumerate(["Persona", "ICP Fit"], 1):
    c = ws3.cell(row=2, column=col, value=val)
    c.fill = fill(PURPLE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

for ti, th in enumerate(tool_short, 3):
    c = ws3.cell(row=2, column=ti, value=th)
    c.fill = fill(ORANGE)
    c.font = font(bold=True, color=WHITE, size=9)
    c.alignment = align(h="center")
    c.border = border_thin()

c = ws3.cell(row=2, column=10, value="Sequence Order")
c.fill = fill(ORANGE)
c.font = font(bold=True, color=WHITE, size=10)
c.alignment = align(h="center")
c.border = border_thin()

# tools_flags: [①, ②, ③, ④, ⑤, ⑥, ⑦]
personas_matrix = [
    ("CHRO / VP HR",                    "Primary Buyer",   [1, 0, 0, 0, 1, 0, 1], "① → ⑤ → ⑦",          WHITE),
    ("L&D / Talent Dev Head",           "Champion",        [0, 0, 0, 1, 1, 1, 1], "④ → ⑥ → ⑤ → ⑦",      GREY),
    ("Talent Acquisition Leader",       "Champion",        [1, 1, 1, 0, 0, 0, 0], "② → ③ → ①",           WHITE),
    ("Workforce Planning / Analytics",  "Influencer",      [1, 0, 1, 0, 1, 0, 1], "① → ③ → ⑤ → ⑦",      GREY),
    ("HR Tech Evaluators & Analysts",   "Amplifier",       [1, 1, 1, 1, 1, 1, 1], "All 7 (content-led)", WHITE),
]

CHECK = "✓"
DASH  = "—"

for ri, (persona, fit, flags, order, bg) in enumerate(personas_matrix, 3):
    ws3.row_dimensions[ri].height = 26
    for col, val in enumerate([persona, fit], 1):
        c = ws3.cell(row=ri, column=col, value=val)
        c.fill = fill(PURPLE_LITE if col == 1 else bg)
        c.font = font(bold=(col == 1), color=TEXT_DARK, size=10)
        c.alignment = align()
        c.border = border_thin()
    for ti, flag in enumerate(flags, 3):
        c = ws3.cell(row=ri, column=ti, value=CHECK if flag else DASH)
        c.fill = fill(GREEN_LITE if flag else bg)
        c.font = font(bold=flag, color=GREEN if flag else TEXT_MUTED, size=12)
        c.alignment = align(h="center")
        c.border = border_thin()
    c = ws3.cell(row=ri, column=10, value=order)
    c.fill = fill(ORANGE_LITE)
    c.font = font(bold=True, color=ORANGE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

for col, w in zip(range(1, 11), [30, 16, 11, 11, 11, 11, 11, 11, 11, 26]):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — OUTREACH PLAYBOOKS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("④ Outreach Playbooks")

ws4.merge_cells("A1:F1")
c = ws4["A1"]
c.value = "PROJECT ARIA — Persona Outreach Playbooks  ·  One Tool Per Message. Always."
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws4.row_dimensions[1].height = 30

ws4.merge_cells("A2:F2")
c = ws4["A2"]
c.value = ("CORE RULE: Never send multiple tool links in one message. "
           "Each message leads with one problem → one tool → earns the right to send the next message based on what the persona does.")
c.fill = fill(ORANGE_LITE)
c.font = font(bold=True, color=ORANGE, size=10)
c.alignment = align(h="left")
c.border = border_thin()
ws4.row_dimensions[2].height = 28

headers4 = ["Persona", "Touch", "Trigger", "Tool", "Message Hook (Opening Line)", "Timing"]
ws4.row_dimensions[3].height = 20
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=3, column=col, value=h)
    c.fill = fill(ORANGE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

playbook_data = [
    # ── CHRO / VP HR ──────────────────────────────────────────────────────────
    ("CHRO / VP HR", "Touch 1", "Enrol Day 0",
     "① AI Automation Index",
     "We built a free tool that takes any job description and returns a scored breakdown of which tasks AI can handle today vs. which still need a human. Takes 60 seconds.",
     "Day 0"),
    ("", "Touch 2A\n(Engaged)", "Used Tool ①\n→ 3-5 days",
     "⑤ Career Path Intelligence",
     "Since you've seen the automation risk picture for [role] — the forward-looking question CHROs ask us next: what career paths remain viable as AI handles more of the work? Paste any skill set and target role. Returns 3 AI-safe paths in 2 minutes.",
     "Day 3-5"),
    ("", "Touch 2B\n(Cold)", "No engage\n→ 10 days",
     "⑦ Emerging Skills Radar",
     "As AI reshapes roles, the skills that matter in 2027 are already appearing in the wild — before any framework has named them. We built a tool that scans real signals and surfaces what's rising for any role or domain. Takes 90 seconds.",
     "Day 10"),
    ("", "Touch 3\n(Sales CTA)", "Used 2 tools\nor Touch 2B",
     "→ Book call",
     "You've now seen both the automation risk picture and the emerging skills signal for [company's roles]. Most CHROs at this point want to see how this runs across their full workforce automatically. Worth 20 minutes?",
     "Day 7-12"),

    # ── L&D / Talent Dev ──────────────────────────────────────────────────────
    ("L&D / Talent Dev Head", "Touch 1", "Enrol Day 0",
     "④ Certification Skill Mapper",
     "We built a tool that takes any list of certifications and maps exactly which skills they prove — with proficiency level and expiry signal. Useful when you're trying to understand what your team actually has validated vs. claimed.",
     "Day 0"),
    ("", "Touch 2A\n(Engaged)", "Used Tool ④\n→ 3-5 days",
     "⑥ Learning Path Generator",
     "Once you know the validated skills — the next question is: what's the fastest path to close the gaps? Our Learning Path tool takes a skill gap and returns a prioritised plan with content type, estimated skill gain, and ROI projection per item.",
     "Day 3-5"),
    ("", "Touch 2B\n(Cold)", "No engage\n→ 10 days",
     "⑤ Career Path Intelligence",
     "Different angle — if you have people wondering where their careers go as AI changes their role, this tool gives them a concrete answer in 2 minutes. Paste current skills + target role. Returns 3 AI-safe paths with skill milestones.",
     "Day 10"),
    ("", "Touch 3A\n(Sales CTA)", "Used 2 tools",
     "→ Book call",
     "You've seen the certification picture and the learning plan logic. What iMocha does is run this continuously across your entire workforce — every employee, every gap, automatically prioritised. Worth a 20-minute conversation?",
     "Day 8-12"),
    ("", "Touch 3B\n(Cold)", "1 or 0 tools\nafter Touch 2",
     "⑦ Emerging Skills Radar",
     "Last one — we built a tool that surfaces skills rising in the wild for any role before any framework captures them. Useful if you're planning curriculum for 2026-2027. 90 seconds, no login.",
     "Day 20"),

    # ── TA Leader / Recruiter ─────────────────────────────────────────────────
    ("TA Leader /\nRecruiter", "Touch 1", "Enrol Day 0",
     "② Resume Skill Extractor",
     "We built a free tool that takes any resume (paste or upload) and returns a structured skill profile — every skill extracted, proficiency-rated, and confidence-scored in seconds. Cuts manual resume screening time for any role.",
     "Day 0"),
    ("", "Touch 2A\n(Engaged)", "Used Tool ②\n→ 3-5 days",
     "③ Skill Normalizer",
     "Related problem — when you're comparing 50 resumes, the same skill appears under 30 different names. Our Skill Normalizer takes any list of skill strings and returns the canonical version of each, deduplicated and framework-mapped. Takes 20 seconds.",
     "Day 3-5"),
    ("", "Touch 2B\n(Cold)", "No engage\n→ 10 days",
     "③ Skill Normalizer",
     "If you've ever had two candidates list 'ML' and 'Machine Learning' and not realised they meant the same thing — we built a tool that solves this. Paste any skill list from a JD or resume, get canonical names back instantly.",
     "Day 10"),
    ("", "Touch 3", "Any path",
     "① AI Automation Index",
     "Beyond extracting and normalising skills — thought you'd want to see which skills in your JDs are being automated. The AI Automation Index scores every task in a role. Useful before you finalise a hiring brief.",
     "Day 15"),
    ("", "Touch 4\n(Sales CTA)", "Used 2+ tools",
     "→ Book call",
     "iMocha does all of this automatically across every open role in your ATS — resume parsing, skill normalisation, and JD scoring before a req goes live. Worth 20 minutes?",
     "Day 18-22"),

    # ── Workforce Planning / Analytics ────────────────────────────────────────
    ("Workforce Planning /\nPeople Analytics", "Touch 1", "Enrol Day 0",
     "① AI Automation Index",
     "We built a tool that takes any job description and returns a task-by-task automation risk score — every task rated 0-100 with a specific AI tool for each. Useful input for your workforce planning models.",
     "Day 0"),
    ("", "Touch 2A\n(Engaged)", "Used Tool ①\n→ 3-5 days",
     "③ Skill Normalizer",
     "For the role you scored — thought you'd want to see the skills side normalised too. Our Skill Normalizer takes raw skill strings from any JD or profile and returns canonical names, deduplicated and framework-mapped. Useful when you're aggregating across roles.",
     "Day 3-5"),
    ("", "Touch 2B\n(Cold)", "No engage\n→ 10 days",
     "⑦ Emerging Skills Radar",
     "As you model future workforce needs — the forward signal is which skills are rising before any framework captures them. We built a radar for this: enter any role or domain, get the emerging skills picture in 90 seconds.",
     "Day 10"),
    ("", "Touch 3\n(Sales CTA)", "Any path",
     "⑤ Career Path Intelligence",
     "One more angle — our Career Path tool takes any skill profile + target role and returns 3 AI-safe paths with skill milestones. Useful for internal mobility modelling. And iMocha runs this logic automatically across your entire workforce.",
     "Day 20"),

    # ── HR Tech Evaluator / Analyst ───────────────────────────────────────────
    ("HR Tech Evaluator /\nAnalyst", "Touch 1", "Enrol Day 0",
     "① AI Automation Index",
     "We've built a portfolio of free AI tools that each demonstrate a specific agent from iMocha's 23-agent skills intelligence architecture. The first is live. Thought it might be useful reference material for your coverage of AI's impact on HR tech.",
     "Day 0"),
    ("", "Touch 2+", "Every new\ntool ships",
     "New tool\n(as released)",
     "We just shipped [Tool Name] — a free tool that [aha moment in one line]. It demonstrates [Agent #X] from iMocha's architecture. Adding it to the portfolio — happy to walk through the full set if useful for anything you're writing or evaluating.",
     "Every 10 days"),
]

persona_colors = {
    "CHRO / VP HR":                     PURPLE_LITE,
    "L&D / Talent Dev Head":            GREEN_LITE,
    "TA Leader /\nRecruiter":           ORANGE_LITE,
    "Workforce Planning /\nPeople Analytics": AMBER_LITE,
    "HR Tech Evaluator /\nAnalyst":     RED_LITE,
}
touch_colors = {
    "Touch 1":         (PURPLE, WHITE),
    "Touch 2A\n(Engaged)": (GREEN, WHITE),
    "Touch 2B\n(Cold)":    (AMBER, WHITE),
    "Touch 3":         (ORANGE, WHITE),
    "Touch 3A\n(Sales CTA)": (ORANGE, WHITE),
    "Touch 3B\n(Cold)":    (TEXT_MUTED, WHITE),
    "Touch 4\n(Sales CTA)": (ORANGE, WHITE),
    "Touch 2+":        (GREEN, WHITE),
}

current_persona = ""
current_persona_bg = WHITE

for ri, (persona, touch, trigger, tool, hook, timing) in enumerate(playbook_data, 4):
    ws4.row_dimensions[ri].height = 54
    if persona:
        current_persona = persona
        current_persona_bg = persona_colors.get(persona, PURPLE_LITE)

    pc = ws4.cell(row=ri, column=1, value=persona if persona else "")
    pc.fill = fill(current_persona_bg)
    pc.font = font(bold=True, color=TEXT_DARK, size=10)
    pc.alignment = align(h="center")
    pc.border = border_thin()

    tc_bg, tc_fg = touch_colors.get(touch, (PURPLE_LITE, TEXT_DARK))
    tc = ws4.cell(row=ri, column=2, value=touch)
    tc.fill = fill(tc_bg)
    tc.font = font(bold=True, color=tc_fg, size=9)
    tc.alignment = align(h="center")
    tc.border = border_thin()

    trig = ws4.cell(row=ri, column=3, value=trigger)
    trig.fill = fill(GREY)
    trig.font = font(color=TEXT_MID, size=9, italic=True)
    trig.alignment = align(h="center")
    trig.border = border_thin()

    toolc = ws4.cell(row=ri, column=4, value=tool)
    toolc.fill = fill(ORANGE_LITE)
    toolc.font = font(bold=True, color=ORANGE, size=10)
    toolc.alignment = align(h="center")
    toolc.border = border_thin()

    hookc = ws4.cell(row=ri, column=5, value=hook)
    hookc.fill = fill(WHITE)
    hookc.font = font(color=TEXT_DARK, size=10)
    hookc.alignment = align()
    hookc.border = border_thin()

    timec = ws4.cell(row=ri, column=6, value=timing)
    timec.fill = fill(PURPLE_LITE)
    timec.font = font(bold=True, color=PURPLE, size=9)
    timec.alignment = align(h="center")
    timec.border = border_thin()

for col, w in zip(range(1, 7), [22, 14, 18, 22, 72, 14]):
    ws4.column_dimensions[get_column_letter(col)].width = w
ws4.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SEQUENCE RULES
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⑤ Sequence Rules")

ws5.merge_cells("A1:D1")
c = ws5["A1"]
c.value = "PROJECT ARIA — Outreach Sequence Trigger Rules  ·  Automation Logic"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws5.row_dimensions[1].height = 30

ws5.merge_cells("A2:D2")
c = ws5["A2"]
c.value = "THREE PATH TYPES"
c.fill = fill(ORANGE)
c.font = font(bold=True, color=WHITE, size=11)
c.alignment = align(h="center")
ws5.row_dimensions[2].height = 20

paths = [
    ("TYPE A — Engaged Path", "Persona uses a tool", "Reference what they found → send next tool as natural follow-on → after 2 tools used → Sales CTA immediately", GREEN_LITE, GREEN),
    ("TYPE B — Cold Path", "Persona opens email but doesn't use tool", "Try different angle → different tool, different problem framing → after 3 touchpoints with no usage → low-intent flag", AMBER_LITE, AMBER),
    ("TYPE C — Champion Path", "Persona uses 2+ tools or shares with colleague", "Auto-trigger Slack alert to AE with enriched profile → AE reaches out within 24h → skip remaining touches", ORANGE_LITE, ORANGE),
]
for ri, (ptype, trigger, action, bg, tc) in enumerate(paths, 3):
    ws5.row_dimensions[ri].height = 40
    for col, val in enumerate([ptype, trigger, action], 1):
        c = ws5.cell(row=ri, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=(col == 1), color=tc if col == 1 else TEXT_DARK, size=10)
        c.alignment = align()
        c.border = border_thin()

ws5.append([])  # spacer

ws5.merge_cells("A7:D7")
c = ws5["A7"]
c.value = "TRIGGER RULES TABLE"
c.fill = fill(ORANGE)
c.font = font(bold=True, color=WHITE, size=11)
c.alignment = align(h="center")
ws5.row_dimensions[7].height = 20

rule_headers = ["Trigger", "Action", "Timing", "Who Acts"]
ws5.row_dimensions[8].height = 20
for col, h in enumerate(rule_headers, 1):
    c = ws5.cell(row=8, column=col, value=h)
    c.fill = fill(PURPLE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

rules = [
    ("Persona completes a tool analysis",           "Enrich in Clay → log in HubSpot → start next touch timer",             "Immediate",    "Automated"),
    ("Persona uses 1 tool",                          "Queue Touch 2A (engaged path) with contextual follow-on tool",          "3-5 days",     "Automated"),
    ("Persona opens email but doesn't use tool",     "Queue Touch 2B (cold path) with different tool, different framing",     "10 days",      "Automated"),
    ("Persona uses 2 tools",                         "Slack alert to AE with enriched profile → remove from outreach sequence","Immediate",   "System → AE"),
    ("AE books meeting",                             "Close sequence, mark as Sales Accepted Lead in HubSpot",                "Immediate",    "AE"),
    ("3 touches sent, zero tool usage",              "Mark low-intent · pause sequence · re-enrol when next tool ships",      "Day 20",       "Automated"),
    ("Persona shares tool (referral URL detected)",  "Enrol colleague in relevant sequence · Alert AE",                       "Immediate",    "System → AE"),
    ("Analyst / evaluator uses 3+ tools",            "Offer full portfolio walkthrough — no sales pressure, PM reaches out",  "3-5 days",     "PM manually"),
]

rule_bgs = [WHITE, GREY, WHITE, GREEN_LITE, GREEN_LITE, AMBER_LITE, PURPLE_LITE, ORANGE_LITE]
for ri, ((trigger, action, timing, who), bg) in enumerate(zip(rules, rule_bgs), 9):
    ws5.row_dimensions[ri].height = 36
    for col, val in enumerate([trigger, action, timing, who], 1):
        c = ws5.cell(row=ri, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(color=TEXT_DARK, size=10)
        c.alignment = align()
        c.border = border_thin()

for col, w in zip(range(1, 5), [44, 56, 14, 18]):
    ws5.column_dimensions[get_column_letter(col)].width = w
ws5.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — METRICS
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("⑥ Metrics")

ws6.merge_cells("A1:E1")
c = ws6["A1"]
c.value = "PROJECT ARIA — Success Metrics & 90-Day Targets"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws6.row_dimensions[1].height = 30

metric_headers = ["Metric", "Category", "30-Day Target", "60-Day Target", "90-Day Target"]
ws6.row_dimensions[2].height = 22
for col, h in enumerate(metric_headers, 1):
    c = ws6.cell(row=2, column=col, value=h)
    c.fill = fill(ORANGE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

metrics = [
    ("Tools live",                         "Build",   "1",     "5",       "7"),
    ("Persona outreach sent",              "Reach",   "500",   "2,500",   "5,000"),
    ("Tool trials completed",              "Reach",   "200",   "1,500",   "5,000"),
    ("Trial-to-Touch-2 conversion rate",   "Reach",   "—",     "25%",     "30%"),
    ("Qualified leads (2+ tools used)",    "Convert", "25",    "150",     "600"),
    ("Sales-accepted leads handed to AE",  "Convert", "10",    "80",      "250"),
    ("Pipeline influenced ($)",            "Revenue", "$100K", "$800K",   "$2M"),
    ("Cost per qualified lead",            "Revenue", "—",     "$12",     "< $10"),
]

cat_colors = {"Build": PURPLE_LITE, "Reach": ORANGE_LITE, "Convert": GREEN_LITE, "Revenue": AMBER_LITE}
cat_text   = {"Build": PURPLE, "Reach": ORANGE, "Convert": GREEN, "Revenue": AMBER}
for ri, (metric, cat, t30, t60, t90) in enumerate(metrics, 3):
    ws6.row_dimensions[ri].height = 24
    mc = ws6.cell(row=ri, column=1, value=metric)
    mc.fill = fill(GREY)
    mc.font = font(color=TEXT_DARK, size=10)
    mc.alignment = align()
    mc.border = border_thin()

    cc = ws6.cell(row=ri, column=2, value=cat)
    cc.fill = fill(cat_colors[cat])
    cc.font = font(bold=True, color=cat_text[cat], size=10)
    cc.alignment = align(h="center")
    cc.border = border_thin()

    for col, val in enumerate([t30, t60, t90], 3):
        c = ws6.cell(row=ri, column=col, value=val)
        c.fill = fill(GREEN_LITE if col == 5 else WHITE)
        c.font = font(bold=(col == 5), color=GREEN if col == 5 else TEXT_DARK, size=10)
        c.alignment = align(h="center")
        c.border = border_thin()

for col, w in zip(range(1, 6), [38, 14, 16, 16, 16]):
    ws6.column_dimensions[get_column_letter(col)].width = w
ws6.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — BUDGET
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("⑦ Budget")

ws7.merge_cells("A1:D1")
c = ws7["A1"]
c.value = "PROJECT ARIA — Budget & Resource Plan  ·  90-Day Pilot"
c.fill = fill(PURPLE)
c.font = font(bold=True, color=WHITE, size=13)
c.alignment = align(h="center")
ws7.row_dimensions[1].height = 30

budget_headers = ["Resource", "Detail", "Monthly Cost", "90-Day Total"]
ws7.row_dimensions[2].height = 22
for col, h in enumerate(budget_headers, 1):
    c = ws7.cell(row=2, column=col, value=h)
    c.fill = fill(ORANGE)
    c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align(h="center")
    c.border = border_thin()

budget_rows = [
    ("PM Time",              "50% of Growth PM's time — existing headcount",         "Internal",  "Internal"),
    ("Technical Intern",     "1 FTE, full-time for 90 days (7 tools in 7 sprints)",  "—",         "~$6,000"),
    ("AI API Costs",         "Claude (Anthropic) + LangSmith per tool run",           "~$800",     "~$2,400"),
    ("Outreach Tools",       "Clay + Apollo licences",                                "~$600",     "~$1,800"),
    ("Paid Amplification",   "LinkedIn Sponsored content + community sponsorships",   "~$1,000",   "~$3,000"),
    ("Contingency (10%)",    "Buffer for tool infrastructure, overruns",              "—",         "~$1,300"),
]

for ri, (res, detail, monthly, total) in enumerate(budget_rows, 3):
    bg = WHITE if ri % 2 == 1 else GREY
    ws7.row_dimensions[ri].height = 28
    for col, val in enumerate([res, detail, monthly, total], 1):
        c = ws7.cell(row=ri, column=col, value=val)
        c.fill = fill(PURPLE_LITE if col == 1 else bg)
        c.font = font(bold=(col == 1), color=TEXT_DARK, size=10)
        c.alignment = align()
        c.border = border_thin()

ws7.row_dimensions[9].height = 28
# Write values first, then merge
for col, val in enumerate(["TOTAL 90-DAY BUDGET", "", "", "~$14,500"], 1):
    c = ws7.cell(row=9, column=col, value=val)
    c.fill = fill(PURPLE if col != 4 else GREEN_LITE)
    c.font = font(bold=True, color=WHITE if col != 4 else GREEN, size=12)
    c.alignment = align(h="center")
    c.border = border_thin()
ws7.merge_cells("A9:C9")

ws7.append([])

write_section_title(ws7, 11, 1, "PROJECTED RETURN ON $14,500", colspan=4, bg=ORANGE)

roi_rows = [
    ("Pipeline Influenced (90-Day)", "$2,000,000",  "138x budget"),
    ("Sales-Accepted Leads",         "250 leads",   "$58 per SAL"),
    ("Tool Trials Completed",        "5,000",       "$2.90 per trial"),
    ("Qualified Leads (2+ tools)",   "600 leads",   "$24 per qualified lead"),
]

for ri, (metric, value, note) in enumerate(roi_rows, 12):
    ws7.row_dimensions[ri].height = 24
    mc = ws7.cell(row=ri, column=1, value=metric)
    mc.fill = fill(GREY)
    mc.font = font(color=TEXT_DARK, size=10)
    mc.alignment = align()
    mc.border = border_thin()
    vc = ws7.cell(row=ri, column=2, value=value)
    vc.fill = fill(GREEN_LITE)
    vc.font = font(bold=True, color=GREEN, size=12)
    vc.alignment = align(h="center")
    vc.border = border_thin()
    nc = ws7.cell(row=ri, column=3, value=note)
    nc.fill = fill(WHITE)
    nc.font = font(color=TEXT_MID, size=10, italic=True)
    nc.alignment = align()
    nc.border = border_thin()

for col, w in zip(range(1, 5), [30, 30, 20, 16]):
    ws7.column_dimensions[get_column_letter(col)].width = w
ws7.sheet_view.showGridLines = False


# ── Save ───────────────────────────────────────────────────────────────────────
output_path = "/Users/adityathakkur/Desktop/ai-automation-index/ARIA_Initiative_Framework.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
